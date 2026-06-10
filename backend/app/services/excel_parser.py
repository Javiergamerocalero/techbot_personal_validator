"""Parseo + validación de filas del Excel de empleados.

Se hace en memoria con `openpyxl` (read-only mode). Si el archivo es
gigantesco (> 50k filas) podemos pasar a streaming, pero para San
Fernando el orden de magnitud es 1k-10k.
"""
import io
from typing import IO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError

from app.models.employee import DocumentType, StatusReason
from app.schemas.employee import (
    EmployeeImportRow,
    ImportError as ImportErrorSchema,
)


HEADERS = [
    "employee_code",
    "document_number",
    "document_type",
    "full_name",
    "status",
    "status_reason",
    "cost_center",
    "tenant_name",
]

_DOC_TYPES = [t.value for t in DocumentType]
_STATUS_REASONS = [r.value for r in StatusReason]


def build_template_xlsx() -> bytes:
    """Genera el `.xlsx` plantilla — headers + hoja de catálogos.

    La hoja "Catalogos" muestra los valores válidos de los enums.
    En la hoja principal hay data validation que limita las celdas
    correspondientes a esos valores.
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Empleados"

    # Headers con formato
    header_fill = PatternFill(
        start_color="222B57", end_color="222B57", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            len(header) + 4, 18
        )

    # Hoja de catálogos
    cat_ws = wb.create_sheet(title="Catalogos")
    cat_ws.cell(row=1, column=1, value="document_type").font = Font(bold=True)
    cat_ws.cell(row=1, column=2, value="status_reason").font = Font(bold=True)
    for i, v in enumerate(_DOC_TYPES, start=2):
        cat_ws.cell(row=i, column=1, value=v)
    for i, v in enumerate(_STATUS_REASONS, start=2):
        cat_ws.cell(row=i, column=2, value=v)

    # Data validation en columnas C (document_type) y F (status_reason)
    dv_doctype = DataValidation(
        type="list", formula1=f'"{",".join(_DOC_TYPES)}"', allow_blank=True
    )
    dv_doctype.add(f"C2:C1048576")
    ws.add_data_validation(dv_doctype)

    dv_reason = DataValidation(
        type="list",
        formula1=f'"{",".join(_STATUS_REASONS)}"',
        allow_blank=True,
    )
    dv_reason.add(f"F2:F1048576")
    ws.add_data_validation(dv_reason)

    # Fila de ejemplo
    ws.append(
        [
            "SF000001",
            "12345678",
            "DNI",
            "Juan Perez Lopez",
            True,
            "ACTIVE",
            "PLANTA CHORRILLOS",
            "San Fernando",
        ]
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def parse_xlsx(
    file_bytes: bytes | IO[bytes],
) -> tuple[list[EmployeeImportRow], list[ImportErrorSchema]]:
    """Lee un .xlsx y devuelve (filas válidas, errores).

    El caller decide qué hacer con las filas válidas (insertar / upsert).
    Las filas con error se reportan al frontend para corrección.
    """
    if isinstance(file_bytes, (bytes, bytearray)):
        stream: IO[bytes] = io.BytesIO(file_bytes)
    else:
        stream = file_bytes
    try:
        wb = load_workbook(stream, read_only=True, data_only=True)
    except Exception as e:
        return (
            [],
            [
                ImportErrorSchema(
                    row=0,
                    column=None,
                    reason=f"No se pudo abrir el archivo: {e}",
                )
            ],
        )

    ws = wb.active
    if ws is None:
        return [], [
            ImportErrorSchema(row=0, column=None, reason="Hoja vacía")
        ]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], [
            ImportErrorSchema(row=0, column=None, reason="Archivo vacío")
        ]

    header_map: dict[str, int] = {}
    for idx, h in enumerate(header_row):
        if h is None:
            continue
        normalized = str(h).strip().lower()
        if normalized in HEADERS:
            header_map[normalized] = idx

    missing = [h for h in HEADERS if h not in header_map]
    if missing:
        return [], [
            ImportErrorSchema(
                row=1,
                column=", ".join(missing),
                reason=f"Faltan columnas: {', '.join(missing)}",
            )
        ]

    valid: list[EmployeeImportRow] = []
    errors: list[ImportErrorSchema] = []

    for row_num, row in enumerate(rows_iter, start=2):
        if row is None or all(c in (None, "") for c in row):
            continue  # fila vacía
        try:
            data = {
                "employee_code": _str_or_none(row[header_map["employee_code"]]),
                "document_number": _str_or_none(
                    row[header_map["document_number"]]
                ),
                "document_type": _str_or_default(
                    row[header_map["document_type"]], "DNI"
                ),
                "full_name": _str_or_none(row[header_map["full_name"]]),
                "status": _to_bool(row[header_map["status"]]),
                "status_reason": _str_or_default(
                    row[header_map["status_reason"]], "ACTIVE"
                ),
                "cost_center": _str_or_none(row[header_map["cost_center"]]),
                "tenant_name": _str_or_none(row[header_map["tenant_name"]]),
            }
            parsed = EmployeeImportRow.model_validate(data)
            valid.append(parsed)
        except ValidationError as ve:
            for err in ve.errors():
                loc = ".".join(str(p) for p in err["loc"]) or None
                errors.append(
                    ImportErrorSchema(
                        row=row_num, column=loc, reason=err["msg"]
                    )
                )
        except Exception as e:  # noqa: BLE001
            errors.append(
                ImportErrorSchema(
                    row=row_num,
                    column=None,
                    reason=f"Error inesperado: {e}",
                )
            )

    return valid, errors


def _str_or_none(v) -> str | None:  # noqa: ANN001
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _str_or_default(v, default: str) -> str:  # noqa: ANN001
    s = _str_or_none(v)
    return s if s is not None else default


def _to_bool(v) -> bool:  # noqa: ANN001
    if isinstance(v, bool):
        return v
    if v is None:
        return True
    if isinstance(v, (int, float)):
        return bool(v)
    s = str(v).strip().lower()
    if s in {"1", "true", "yes", "y", "si", "sí", "x"}:
        return True
    if s in {"0", "false", "no", "n", ""}:
        return False
    raise ValueError(f"Valor de status no reconocido: {v!r}")
